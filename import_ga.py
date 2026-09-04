import openpyxl
import pymysql
import sys

def main():
    print("Connecting to MariaDB at 10.9.9.110...")
    try:
        conn = pymysql.connect(
            host='10.9.9.110',
            port=3306,
            user='userdb',
            password='sahabat25*',
            database='asset_inventory',
            autocommit=True
        )
    except Exception as e:
        print(f"Error connecting to DB: {e}")
        return

    cursor = conn.cursor()

    # 1. Setup columns & tables
    ddls = [
        "ALTER TABLE assets ADD COLUMN legacy_inv_code VARCHAR(100) NULL",
        "ALTER TABLE assets ADD COLUMN sticker_status VARCHAR(50) DEFAULT 'UNKNOWN'",
        """CREATE TABLE IF NOT EXISTS asset_maintenance_logs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_id VARCHAR(100) NOT NULL,
            type VARCHAR(50) NOT NULL,
            performed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            performed_by VARCHAR(100) NOT NULL,
            notes TEXT,
            INDEX idx_asset_maint (asset_id)
        )""",
        """CREATE TABLE IF NOT EXISTS asset_part_replacements (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_id VARCHAR(100) NOT NULL,
            part_name VARCHAR(100) NOT NULL,
            action_type VARCHAR(50) NOT NULL,
            old_spec VARCHAR(255),
            new_spec VARCHAR(255),
            reason TEXT,
            replaced_at DATETIME NOT NULL,
            technician VARCHAR(100) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_asset_parts (asset_id)
        )""",
        """CREATE TABLE IF NOT EXISTS asset_location_history (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_id VARCHAR(100) NOT NULL,
            from_location VARCHAR(100),
            to_location VARCHAR(100),
            moved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            moved_by VARCHAR(100) NOT NULL,
            reason TEXT,
            INDEX idx_asset_loc (asset_id)
        )"""
    ]

    for ddl in ddls:
        try:
            cursor.execute(ddl)
        except Exception as e:
            pass # Ignore duplicate column errors

    print("Loading INVENTARIS PERANGKAT IT.xlsx...")
    wb = openpyxl.load_workbook('INVENTARIS PERANGKAT IT.xlsx', data_only=True)

    categories = [
        ("Monitor", "MN", "MONITOR - NO INV", "MONITOR - INVENTARIS"),
        ("CPU", "PC", "CPU - NO INV", "CPU - INVENTARIS"),
        ("Keyboard", "KB", "KEYBOARD - NO INV", "KEYBOARD - INVENTARIS"),
        ("Mouse", "MS", "MOUSE - NO INV", "MOUSE - INVENTARIS"),
        ("Headset", "HD", "HEADSET - NO INV", "HEADSET - INVENTARIS"),
    ]

    total_inserted = 0
    total_updated = 0

    for type_name, prefix, no_inv_sheet, inv_sheet in categories:
        print(f"\nProcessing {type_name} ({prefix})...")

        # 1. Process NO INV sheet
        if no_inv_sheet in wb.sheetnames:
            sheet = wb[no_inv_sheet]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 3:
                for r_idx, row in enumerate(rows[3:], start=4):
                    if not row or len(row) < 3 or row[2] is None:
                        continue
                    raw_code = str(row[2]).strip()
                    if not raw_code:
                        continue

                    # Asset ID e.g. MN-0181
                    parts = raw_code.split('/')
                    asset_id = raw_code
                    if len(parts) >= 2:
                        asset_id = f"{parts[0]}-{parts[1]}"

                    stiker_val = str(row[3]).upper() if len(row) > 3 and row[3] is not None else ""
                    sticker_status = "STICKERED" if (stiker_val == "TRUE" or stiker_val == "1" or stiker_val == "YES") else "UNSTICKERED"

                    note = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

                    sql = """
                        INSERT INTO assets (id, type, status, location, specs, note, legacy_inv_code, sticker_status)
                        VALUES (%s, %s, 'AVAILABLE', 'Ruang IT', 'Source: GA Master NO INV', %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                            legacy_inv_code = VALUES(legacy_inv_code),
                            sticker_status = VALUES(sticker_status),
                            note = COALESCE(NULLIF(VALUES(note), ''), note)
                    """
                    cursor.execute(sql, (asset_id, type_name, note, raw_code, sticker_status))
                    total_inserted += 1

        # 2. Process INVENTARIS sheet
        if inv_sheet in wb.sheetnames:
            sheet = wb[inv_sheet]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r_idx, row in enumerate(rows[1:], start=2):
                    if not row or len(row) < 3 or row[2] is None:
                        continue
                    no_inv_raw = str(row[2]).strip()
                    if not no_inv_raw:
                        continue
                    try:
                        no_num = int(float(no_inv_raw))
                        no_str = f"{no_num:04d}"
                    except ValueError:
                        no_str = no_inv_raw.zfill(4)

                    asset_id = f"{prefix}-{no_str}"

                    # Status: TERPAKAI (col 3), RUSAK (col 4), PERBAIKI (col 5)
                    terpakai = str(row[3]).upper() if len(row) > 3 and row[3] is not None else ""
                    rusak = str(row[4]).upper() if len(row) > 4 and row[4] is not None else ""
                    perbaiki = str(row[5]).upper() if len(row) > 5 and row[5] is not None else ""

                    status = "AVAILABLE"
                    if rusak in ["TRUE", "1", "YES", "YA"]:
                        status = "BROKEN"
                    elif perbaiki in ["TRUE", "1", "YES", "YA"]:
                        status = "REPAIRING"
                    elif terpakai in ["TRUE", "1", "YES", "YA"]:
                        status = "IN_USE"

                    note = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

                    sql = """
                        INSERT INTO assets (id, type, status, location, specs, note, sticker_status)
                        VALUES (%s, %s, %s, 'Ruang IT', 'Source: GA Master INVENTARIS', %s, 'STICKERED')
                        ON DUPLICATE KEY UPDATE
                            status = VALUES(status),
                            note = COALESCE(NULLIF(VALUES(note), ''), note)
                    """
                    cursor.execute(sql, (asset_id, type_name, status, note))
                    total_updated += 1

    print(f"\nImport Completed Successfully!")
    print(f"Total processed/inserted records: {total_inserted}")
    print(f"Total status updates from INVENTARIS sheets: {total_updated}")

    # Summary count by type
    cursor.execute("SELECT type, COUNT(*) FROM assets GROUP BY type")
    print("\nCurrent Database Asset Summary:")
    for t_name, count in cursor.fetchall():
        print(f"  - {t_name}: {count} assets")

if __name__ == "__main__":
    main()
